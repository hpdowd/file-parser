#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = ["pdfplumber>=0.11", "openpyxl>=3.1", "cryptography>=42"]
# ///
"""
parse_incidents.py
==================
Extract the marked ("***") cells from structured incident-report PDF pages
and condense many pages into a single, easy-to-read output file (Excel .xlsx by
default; CSV or JSON optional).

The report is a fixed, software-generated template, so every
page has an identical layout. Extraction is *anchor based*: each field is located
by a fixed label on the page (e.g. "Local Station:", "Narrative", "Stolen") and
the value is read from the cell relative to that label. This means it keeps
working on real pages where the cells hold real data instead of "***", and it
naturally handles a Narrative that wraps over several lines (up to ~100 chars).

Usage
-----
    # one PDF (which may contain many pages) -> incidents.xlsx
    uv run parse_incidents.py report.pdf

    # a whole folder of PDFs -> one combined workbook
    uv run parse_incidents.py /path/to/pdfs -o incidents.xlsx

    # other formats
    uv run parse_incidents.py report.pdf -f csv -o incidents.csv
    uv run parse_incidents.py report.pdf -f json

    # show exactly what was extracted from each page (great for new real data)
    uv run parse_incidents.py report.pdf --debug

    # verify the parser against the bundled template (every cell should be "***")
    uv run parse_incidents.py --self-test

Adding more fields later
------------------------
The extracted columns are declared in the FIELDS list near the bottom. To add a
field you append ONE entry: (key, column_header, locator_function). Several small
locator helpers (label_right / cell_below / between / roles_cell) make most new
fields a one-liner. No other code needs to change.
"""
from __future__ import annotations

import argparse
import csv
import getpass
import json
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional

import pdfplumber


# --------------------------------------------------------------------------- #
#  Low-level geometry helpers                                                  #
# --------------------------------------------------------------------------- #
def _norm(s: str) -> str:
    """Collapse runs of whitespace and strip."""
    return re.sub(r"\s+", " ", s or "").strip()


def _same(a: str, b: str) -> bool:
    """Case-insensitive, whitespace-insensitive equality (for banner cross-checks)."""
    return _norm(str(a)).casefold() == _norm(str(b)).casefold()


class PageView:
    """A thin wrapper over a pdfplumber page giving word/line/region lookups.

    All coordinates are pdfplumber's: origin top-left, `top`/`bottom` grow
    downward, in PDF points. We never hard-code coordinates; everything is found
    at run time by matching label text, so minor template shifts don't matter.
    """

    LINE_TOL = 3.0  # words within this vertical distance are on the same line

    def __init__(self, page) -> None:
        self.width = float(page.width)
        self.height = float(page.height)
        self.words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
        for w in self.words:
            w["cx"] = (w["x0"] + w["x1"]) / 2.0
            w["cy"] = (w["top"] + w["bottom"]) / 2.0
        self.lines = self._group_lines()
        # Filled rectangles — used to detect the red cell highlights the report
        # paints on cells that need review (see `_is_red` / loc_supp_role_info).
        self.rects = list(page.rects)

    def _group_lines(self) -> list[dict]:
        ws = sorted(self.words, key=lambda w: (w["top"], w["x0"]))
        groups: list[list[dict]] = []
        for w in ws:
            if groups and abs(w["top"] - groups[-1][-1]["top"]) > self.LINE_TOL:
                groups.append([])
            elif not groups:
                groups.append([])
            groups[-1].append(w)
        lines = []
        for g in groups:
            lines.append(
                {
                    "top": min(x["top"] for x in g),
                    "bottom": max(x["bottom"] for x in g),
                    "words": sorted(g, key=lambda x: x["x0"]),
                }
            )
        return lines

    # -- finding anchors --------------------------------------------------- #
    def find(
        self,
        text: str,
        y_range: Optional[tuple[float, float]] = None,
        occurrence: int = 0,
    ) -> Optional[dict]:
        """Find a contiguous run of words whose joined text == `text`.

        Returns a bbox dict {x0,x1,top,bottom} or None. `y_range` restricts the
        search to lines whose top is within (lo, hi) — used to disambiguate a
        label that appears more than once (e.g. "Incident Type").
        """
        target = _norm(text)
        toks = target.split(" ")
        n = len(toks)
        hits: list[dict] = []
        for ln in self.lines:
            if y_range and not (y_range[0] <= ln["top"] <= y_range[1]):
                continue
            ws = ln["words"]
            for i in range(len(ws) - n + 1):
                run = ws[i : i + n]
                if _norm(" ".join(x["text"] for x in run)) == target:
                    hits.append(
                        {
                            "x0": run[0]["x0"],
                            "x1": run[-1]["x1"],
                            "top": min(x["top"] for x in run),
                            "bottom": max(x["bottom"] for x in run),
                        }
                    )
        if not hits:
            return None
        return hits[min(occurrence, len(hits) - 1)]

    def line_at(self, top: float, tol: float = LINE_TOL) -> Optional[dict]:
        for ln in self.lines:
            if abs(ln["top"] - top) <= tol:
                return ln
        return None

    # -- reading a rectangular region -------------------------------------- #
    def region(
        self, x0: float, x1: float, top: float, bottom: float, sep: str = " "
    ) -> str:
        """Join, in reading order, every word whose centre lies in the rectangle."""
        sel = [
            w
            for w in self.words
            if x0 <= w["cx"] <= x1 and top <= w["cy"] <= bottom
        ]
        sel.sort(key=lambda w: (round(w["top"] / 2.0), w["x0"]))
        return _norm(sep.join(w["text"] for w in sel))


# --------------------------------------------------------------------------- #
#  Field locators                                                             #
#  Each takes a PageView and returns the extracted string ("" if absent).      #
# --------------------------------------------------------------------------- #
def loc_incident_no(p: PageView) -> str:
    return _table_cell(p, "Occurred", "Incident No.", "Incident Type",
                       "Reporting Member Details")


def loc_incident_type(p: PageView) -> str:
    return _table_cell(p, "Occurred", "Incident Type", "Occurred",
                       "Reporting Member Details")


# The report is grouped by three nested black banners, outer -> inner. Each is a
# full-width bar split into [label | value | "N incident(s)" + dept code]. A page
# repeats only the levels that changed since the previous page (so 1, 2 or 3 bars),
# the classic nested group-header pattern. (key, label-on-page, output-column-header)
BANNER_LEVELS = [
    ("grp_why",     "Why on Report?:", "Why on Report"),
    ("grp_type",    "Incident Type:",  "Type Group"),
    ("grp_station", "Local Station:",  "Location"),
]


def _read_banner(p: PageView, label: str) -> Optional[dict]:
    """Read one black banner row identified by `label` (e.g. "Local Station:").
    Returns {'value', 'dept', 'count'} or None if that banner isn't on the page.

    Layout of the row:  <label>   <value …>   <N incident(s)>  <dept>
    The value is bounded on the right by the "N incident(s)" count digits; the dept
    is the highlighted cell after the word "incident(s)". (Generalises the old
    Local-Station-only readers so it works for all three banner levels.)"""
    lbl = p.find(label)
    if not lbl:
        return None
    line = p.line_at(lbl["top"])
    right_words = [w for w in line["words"] if w["x0"] > lbl["x1"]] if line else []
    digits = [w for w in right_words if re.fullmatch(r"\d+", w["text"])]
    value_right = (min(d["x0"] for d in digits) - 5) if digits else (lbl["x1"] + 400)
    value = p.region(lbl["x1"] + 2, value_right, lbl["top"] - 1, lbl["bottom"] + 1)
    inc = [w for w in right_words if re.match(r"incident", w["text"], re.I)]
    dept_left = (max(w["x1"] for w in inc) + 5) if inc else p.width * 0.88
    dept = p.region(dept_left, p.width, lbl["top"] - 1, lbl["bottom"] + 1)
    return {"value": value, "dept": dept, "count": digits[0]["text"] if digits else ""}


def loc_investigating(p: PageView) -> str:
    return _table_cell(p, "Reporting Member Details",
                       "Investigating Member Details",
                       "Nominated Supervisor Details", "Scene (Occurred at)")


def loc_nominated(p: PageView) -> str:
    return _table_cell(p, "Reporting Member Details",
                       "Nominated Supervisor Details", "Reason Included on Report",
                       "Scene (Occurred at)")


def loc_case_no(p: PageView) -> str:
    # Scene row: the "Case No." column, between "Force Used" and "Incident Outcome".
    return _table_cell(p, "Scene (Occurred at)", "Case No.", "Incident Outcome",
                       "Narrative")


def loc_review_station(p: PageView) -> str:
    """The per-incident station, read from the 'Station (Review Station)' column of
    the incident table (the value sits below that header). Unlike the group banner
    this is present on every incident, so it never blanks out on a page where the
    banner isn't repeated.

    It is the right-most column of the header row, so we reuse `_table_cell`: anchor
    on the header line via "Occurred", take the column from "Station (Review
    Station)" rightward to the page edge, and bound the cell above the next section
    ("Reporting Member Details"). `_table_cell` keys off the column's own header, so
    the wrapped second header line ("Due", etc.) — which sits in other columns — is
    excluded automatically.

    Returns the full "<Local Station> (<Review Station>)" cell. parse_files splits it
    with `_split_review_station`: it verifies the Local Station part against the
    grp_station banner, then keeps only the bracketed Review Station for the column."""
    return _table_cell(p, "Occurred", "Station (Review Station)", None,
                       "Reporting Member Details")


def _split_review_station(value: str) -> tuple[str, str]:
    """Split "<Local Station> (<Review Station>)" into (local, review). The local part
    repeats the Local Station banner; the bracketed review part is the per-incident
    value we keep. No brackets -> (whole value, "")."""
    inner = re.findall(r"\(([^)]*)\)", value)
    if inner:
        return value[:value.rfind("(")].strip(), inner[-1].strip()
    return value, ""


# Patterns scanned WITHIN the Narrative (the full narrative is not exported).
#  - Sec 19: "Sec"/"Section" then optional separators then 19, where 19 is not
#    the start of a longer number (so "Sec 192" / "...1994" don't match).
#  - Reference no: F<digits>x<2 digits>, e.g. F38094x24 (the "x" is the 3rd-last
#    character). Case-insensitive; assumes the middle/trailing parts are digits.
SEC19_RE = re.compile(r"sec(?:tion|\.)?[\s.:;,\-]*19(?!\d)", re.IGNORECASE)
REFNUM_RE = re.compile(r"\bF\d+x\d{2}\b", re.IGNORECASE)


def loc_narrative(p: PageView) -> str:
    """Extract the Narrative cell text. Cached on the page because the Sec 19 and
    reference-number fields scan it too (it is no longer a column itself)."""
    cached = getattr(p, "_narrative_cache", None)
    if cached is not None:
        return cached
    n = p.find("Narrative")
    if not n:
        text = ""
    else:
        q = p.find("Quality Control Tests")
        bottom = (q["top"] - 2) if q else (n["bottom"] + 60)
        text = p.region(6, p.width - 6, n["bottom"] + 1, bottom, sep=" ")
    p._narrative_cache = text
    return text


def loc_sec19(p: PageView) -> str:
    """"Yes"/"No" — whether the Narrative references Sec 19."""
    return "Yes" if SEC19_RE.search(loc_narrative(p)) else "No"


def loc_reference_no(p: PageView) -> str:
    """The F…x… reference number(s) found in the Narrative, else "N/A"."""
    seen: list[str] = []
    for m in REFNUM_RE.finditer(loc_narrative(p)):
        if m.group(0) not in seen:
            seen.append(m.group(0))
    return ", ".join(seen) if seen else "N/A"


def _is_red(color) -> bool:
    """True for the report's red cell-highlight fills (it paints cells that need
    review red). Matches both shades seen — tomato (1, .39, .28) and brick
    (.80, .36, .36) — without catching yellow/blue/grey/green fills."""
    if not isinstance(color, (list, tuple)) or len(color) != 3:
        return False
    try:
        r, g, b = (float(color[0]), float(color[1]), float(color[2]))
    except (TypeError, ValueError):
        return False
    return r > 0.55 and g < 0.55 and b < 0.55 and r - g > 0.2 and r - b > 0.2


def loc_supp_role_info(p: PageView) -> str:
    """Supplementary role note. The report paints the review-worthy cells RED, so if
    any red cell is present we return its text (the noteworthy value, joined with
    " | " if several). Otherwise we fall back to the first role row's cell, or "None"
    if that's blank."""
    r = _roles(p)
    if not r or not r["rows"]:
        return ""
    rows = r["rows"]
    reds = [rr for rr in p.rects
            if rr.get("fill") and _is_red(rr.get("non_stroking_color"))]

    def cell(i: int) -> tuple[str, float]:
        top = rows[i]
        bottom = (rows[i + 1] - 2) if i + 1 < len(rows) else (r["end"] - 2)
        return p.region(*r["cols"]["supp"], top - 2, bottom), (top + bottom) / 2.0

    if reds:
        notes: list[str] = []
        for i in range(len(rows)):
            text, mid = cell(i)
            if text and text not in notes \
                    and any(rr["top"] <= mid <= rr["bottom"] for rr in reds):
                notes.append(text)
        if notes:
            return " | ".join(notes)
    # No red cell (or the red ones were blank): first role row's cell, else "None".
    return cell(0)[0] or "None"


def loc_stolen_eur(p: PageView) -> str:
    details = _roles_cell(p, "stolen", "details")
    if not details:
        return ""
    m = re.search(r"€\s*([0-9][0-9.,]*|\*+)", details)
    return m.group(1) if m else ""


# -- shared building blocks used by the locators above --------------------- #
def _table_cell(
    p: PageView,
    header_anchor: str,
    col_left_header: str,
    col_right_header: Optional[str],
    next_section: str,
) -> str:
    """Read a column's value cell from a section table.

    `header_anchor`  - a label unique to the header line (locates the row).
    `col_left_header`/`col_right_header` - the column's own header and the next
        column's header; their left edges bound the column horizontally
        (None right header => extend to the page edge).
    `next_section`   - the label of the following section, bounding the cell.

    The whole cell is captured: every line in this column from just below the
    column header down to the next section. That keeps multi-line values intact
    (e.g. a "Reg# - name (Unit x)" member-details entry that wraps over two or
    three lines) instead of returning only the last line. Because each column's
    header sits at the top of its own column, using that header's bottom as the
    upper bound also keeps wrapped *header* text (which belongs to other columns)
    out of the value.
    """
    h = p.find(header_anchor)
    if not h:
        return ""
    ywin = (h["top"] - 4, h["bottom"] + 4)
    left = p.find(col_left_header, y_range=ywin)
    if not left:
        return ""
    right = p.find(col_right_header, y_range=ywin) if col_right_header else None
    x_left = left["x0"] - 2
    x_right = (right["x0"] - 2) if right else (p.width - 2)
    ns = p.find(next_section)
    top = left["bottom"] + 1
    bottom = (ns["top"] - 1) if ns else p.height
    if bottom <= top:
        return ""
    return p.region(x_left, x_right, top, bottom)


def _roles(p: PageView) -> Optional[dict]:
    """Map the supplementary "Role / Details / Supplementary Role Information"
    table: returns column x-ranges and the y of each role row."""
    h = p.find("Supplementary Role Information")
    if not h:
        return None
    ywin = (h["top"] - 4, h["bottom"] + 4)
    details = p.find("Details", y_range=ywin)
    role = p.find("Role", y_range=ywin)
    details_x0 = details["x0"] if details else 150.0
    role_x0 = role["x0"] if role else 8.0
    cols = {
        "role": (role_x0 - 24, details_x0 - 2),
        "details": (details_x0 - 2, h["x0"] - 2),
        "supp": (h["x0"] - 2, p.width - 2),
    }
    # Bound the table at the footer. Its wording/position varies between report
    # versions (older pages have a "Produced: … Page x/y" line, newer ones only
    # the confidentiality notice), so anchor on either word. If neither is found
    # (e.g. a sanitized page where they're masked, or a reworded footer), fall back
    # to the bottom 5% so the footer line can't be read as an extra role row.
    footer_tops = [w["top"] for w in p.words
                   if "confidential" in w["text"].lower() or w["text"] == "Produced:"]
    end = min(footer_tops) if footer_tops else p.height * 0.95
    rows = []
    for ln in p.lines:
        if ln["top"] <= h["bottom"] or ln["top"] >= end:
            continue
        if any(cols["role"][0] <= w["cx"] <= cols["role"][1] for w in ln["words"]):
            rows.append(ln["top"])
    return {"cols": cols, "rows": sorted(set(rows)), "end": end}


def _roles_cell(p: PageView, role_prefix: str, column: str) -> str:
    """Text of one cell in the roles table: the row whose Role label starts with
    `role_prefix` (case-insensitive), in column 'role'|'details'|'supp'."""
    r = _roles(p)
    if not r or not r["rows"]:
        return ""
    rows = r["rows"]
    for i, top in enumerate(rows):
        bottom = (rows[i + 1] - 2) if i + 1 < len(rows) else (r["end"] - 2)
        label = p.region(*r["cols"]["role"], top - 2, bottom)
        if label.lower().startswith(role_prefix.lower()):
            return p.region(*r["cols"][column], top - 2, bottom)
    return ""


# --------------------------------------------------------------------------- #
#  Field registry  --  ADD NEW COLUMNS HERE (key, header, locator)            #
# --------------------------------------------------------------------------- #
@dataclass(frozen=True)
class Field:
    key: str
    header: str
    locate: Callable[[PageView], str]


# Order matters: it drives the column order. Incident No. comes first — it is the
# field users scan for most on each incident (it is also emphasised in write_xlsx).
FIELDS: list[Field] = [
    Field("incident_no",          "Incident No.",                   loc_incident_no),
    Field("case_no",              "Case No.",                       loc_case_no),
    Field("incident_type",        "Incident Type",                  loc_incident_type),
    Field("review_station",       "Review Station",                 loc_review_station),
    Field("investigating_member", "Investigating Member",           loc_investigating),
    Field("nominated_supervisor", "Nominated Supervisor",           loc_nominated),
    Field("sec19",                "Sec 19",                         loc_sec19),
    Field("reference_no",         "Ref. No.",                       loc_reference_no),
    Field("supp_role_info",       "Supp. Role",                     loc_supp_role_info),
    Field("stolen_eur",           "Stolen (€)",                     loc_stolen_eur),
]

# Provenance columns are prepended so each row can be traced back to its page.
PROVENANCE = [("source_file", "Source File"), ("page", "Page")]

# Group columns, filled DOWN from the nested black banners rather than read per page
# — each banner heads every incident beneath it until the next one, so these can't
# be plain per-page FIELDS locators. The three levels (see BANNER_LEVELS) plus the
# department code drive both the filled-down columns and the nested print dividers.
GROUP = [(key, header) for key, _label, header in BANNER_LEVELS] + \
        [("grp_dept", "Dept Code")]


# --------------------------------------------------------------------------- #
#  Page -> row                                                                #
# --------------------------------------------------------------------------- #
# Structural section labels found on a full incident page. The labels themselves
# are never redacted (only their values are), so counting them reliably tells a
# real incident page from a bare roles-table overflow on the following page.
_SECTION_ANCHORS = (
    "Incident No.", "Occurred", "Reported", "Reporting Member Details",
    "Scene (Occurred at)", "Narrative", "Quality Control Tests",
    "Supplementary Role Information", "Local Station:",
)


def is_incident_page(pv: PageView) -> bool:
    """True if the page looks like a full incident page rather than a bare
    continuation (a long roles table overflowing onto the next page). Counts the
    structural section labels above and treats two or more as an incident page.
    Using several labels means one bit of wording drift can't drop a real page."""
    return sum(1 for a in _SECTION_ANCHORS if pv.find(a) is not None) >= 2


def extract_banners(pv: PageView) -> dict:
    """Read whichever of the three nested banners (Why on Report? / Incident Type /
    Local Station) appear on this page. Returns {level_key: {'value','dept','count'}}
    for each level present (absent levels omitted). Callers carry each level forward
    until its banner reappears — and because an outer change always re-emits the
    inner bars on the same page, updating per present level keeps the nesting right."""
    out: dict = {}
    for key, label, _ in BANNER_LEVELS:
        b = _read_banner(pv, label)
        if b:
            out[key] = b
    return out


def parse_region_map(text: str) -> dict[str, tuple[str, str]]:
    """Parse a region/station map (Markdown): '# Region' headings, each followed
    by a '> CODE - Name' line and a bullet list of every station in that region
    ('- Station'). Returns {station_name_casefolded: (code, name)} for matching
    against the Local Station banner value. See region_stations.md (local-only,
    gitignored) for the format and real data."""
    mapping: dict[str, tuple[str, str]] = {}
    code = name = ""
    for raw in text.splitlines():
        line = raw.strip()
        if line.startswith("#"):
            code = name = ""
        elif line.startswith(">"):
            m = re.match(r">\s*(\S+)\s*-\s*(.+)$", line)
            if m:
                code, name = m.group(1), m.group(2).strip()
        elif line.startswith("-") and code:
            mapping[_norm(line[1:]).casefold()] = (code, name)
    return mapping


def load_region_map(path: Path) -> dict[str, tuple[str, str]]:
    """Load a PLAINTEXT region map file. Prefer encrypt_region_map/
    decrypt_region_map for anything kept on disk long-term — see main()'s
    --encrypt-region-map, which keeps the real names/codes out of plaintext."""
    return parse_region_map(path.read_text(encoding="utf-8"))


# --------------------------------------------------------------------------- #
#  Region map at-rest encryption — so the real station/region names and codes  #
#  are never sitting in plaintext on disk between runs. Passphrase-based:      #
#  scrypt derives a key from the passphrase + a random salt (stored alongside  #
#  the ciphertext), AES-256-GCM encrypts (authenticated, so a wrong passphrase #
#  or corrupted file fails loudly instead of silently returning garbage).      #
#  File layout: 16-byte salt | 12-byte nonce | ciphertext (GCM tag included).  #
# --------------------------------------------------------------------------- #
def _region_map_key(passphrase: str, salt: bytes) -> bytes:
    from cryptography.hazmat.primitives.kdf.scrypt import Scrypt
    return Scrypt(salt=salt, length=32, n=2**14, r=8, p=1).derive(passphrase.encode())


def encrypt_region_map(src: Path, dest: Path, passphrase: str) -> None:
    """Encrypt the plaintext region map at `src` to `dest`."""
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    salt, nonce = os.urandom(16), os.urandom(12)
    key = _region_map_key(passphrase, salt)
    ciphertext = AESGCM(key).encrypt(nonce, src.read_bytes(), None)
    dest.write_bytes(salt + nonce + ciphertext)


def decrypt_region_map(path: Path, passphrase: str) -> dict[str, tuple[str, str]]:
    """Decrypt an encrypted region map and parse it. Raises ValueError on a wrong
    passphrase or corrupted file (never returns partial/garbage data)."""
    from cryptography.exceptions import InvalidTag
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    data = path.read_bytes()
    salt, nonce, ciphertext = data[:16], data[16:28], data[28:]
    key = _region_map_key(passphrase, salt)
    try:
        plaintext = AESGCM(key).decrypt(nonce, ciphertext, None)
    except InvalidTag:
        raise ValueError("wrong passphrase, or the file is corrupted")
    return parse_region_map(plaintext.decode("utf-8"))


def extract_page(pv: PageView, source: str, page_no: int, debug: bool = False) -> dict:
    row: dict = {"source_file": source, "page": page_no}
    for f in FIELDS:
        try:
            row[f.key] = f.locate(pv)
        except Exception as exc:  # never let one field kill the row
            row[f.key] = ""
            if debug:
                print(f"    ! {f.key}: {exc!r}", file=sys.stderr)
    if debug:
        print(f"  {source} p{page_no}")
        for f in FIELDS:
            print(f"    {f.key:22s} = {row[f.key]!r}")
        # The narrative is scanned (Sec 19 / Reference No.) but not exported; show
        # it here so detection can be checked against the source text.
        print(f"    {'narrative (scanned)':22s} = {loc_narrative(pv)!r}")
    return row


def iter_pdf_paths(inputs: list[str]) -> list[Path]:
    paths: list[Path] = []
    for raw in inputs:
        pth = Path(raw)
        if pth.is_dir():
            paths.extend(sorted(p for p in pth.rglob("*") if p.suffix.lower() == ".pdf"))
        elif pth.is_file() and pth.suffix.lower() == ".pdf":
            paths.append(pth)
        else:
            print(f"warning: skipping {raw!r} (not a PDF file or directory)",
                  file=sys.stderr)
    return paths


def parse_files(paths: list[Path], debug: bool = False,
                region_map: Optional[dict[str, tuple[str, str]]] = None
                ) -> tuple[list[dict], list[dict]]:
    """Returns (rows, skipped). `skipped` lists the continuation/non-incident pages
    that were NOT extracted — {source_file, page, source_path} — so a reviewer can
    open each one (the xlsx 'Skipped Pages' sheet links them) and confirm nothing
    real was dropped.

    `region_map` (see load_region_map) is optional and applies ONLY to Sec 19
    incidents: when the row's Local Station banner matches a mapped station, its
    region code is appended to the banner and the mapped name is filled into
    Investigating Member (which a Sec 19 incident normally leaves blank)."""
    rows: list[dict] = []
    skipped: list[dict] = []
    for pdf_path in paths:
        try:
            with pdfplumber.open(pdf_path) as pdf:
                views: list[tuple[int, PageView]] = []
                for i, page in enumerate(pdf.pages, start=1):
                    try:
                        views.append((i, PageView(page)))
                    except Exception as exc:
                        print(f"warning: failed on {pdf_path.name} page {i}: {exc}",
                              file=sys.stderr)
                flags = [is_incident_page(pv) for _, pv in views]
                # Safety net: never silently drop a whole file. If nothing matched
                # the expected layout, extract every page rather than emit nothing.
                force = bool(views) and not any(flags)
                if force:
                    print(f"warning: {pdf_path.name}: no page matched the expected "
                          f"incident layout — extracting all {len(views)} page(s) "
                          f"anyway. Please share a sample so the parser can be tuned.")
                # Last-seen value (and stated count) of each nested banner level,
                # reset per file so one report can't inherit another's groups. Each
                # level is updated whenever its bar reappears; the dept comes from the
                # innermost (station) bar. Updated on every page — even skipped ones —
                # so a banner on a continuation page still carries forward.
                current = {k: "" for k, _l, _h in BANNER_LEVELS}
                counts = {k: "" for k, _l, _h in BANNER_LEVELS}
                dept = ""
                src_path = str(pdf_path.resolve())  # for the xlsx page hyperlinks
                for (i, pv), flag in zip(views, flags):
                    banners = extract_banners(pv)
                    for key, _l, _h in BANNER_LEVELS:
                        if key in banners:
                            current[key] = banners[key]["value"]
                            counts[key] = banners[key]["count"]
                            if key == "grp_station":
                                dept = banners[key]["dept"]
                    if banners and debug:
                        print(f"  {pdf_path.name} p{i}: banners -> " + ", ".join(
                            f"{k}={banners[k]['value']!r}" for k in banners))
                    if not flag and not force:
                        skipped.append({"source_file": pdf_path.name, "page": i,
                                        "source_path": src_path})
                        if debug:
                            print(f"  {pdf_path.name} p{i}: continuation/non-incident "
                                  "page — skipped")
                        continue
                    try:
                        row = extract_page(pv, pdf_path.name, i, debug)
                        for key, _l, _h in BANNER_LEVELS:
                            row[key] = current[key]
                            row[key + "_n"] = counts[key]  # banner-stated count
                        row["grp_dept"] = dept
                        row["source_path"] = src_path
                        # Cross-check the per-incident values against the banner this
                        # row falls under: Incident Type vs grp_type, and the Review
                        # Station's local-station prefix vs grp_station. Any mismatch
                        # raises the "check" flag; the reason(s) are recorded in
                        # check_detail (used by the Discrepancies sheet + a cell note).
                        # Done while the full review-station cell is available; then
                        # reduce that column to its bracketed Review Station value.
                        local, review = _split_review_station(row.get("review_station", ""))
                        issues = []
                        if (current["grp_type"] and row.get("incident_type")
                                and not _same(row["incident_type"], current["grp_type"])):
                            issues.append("Incident Type differs from banner")
                        if (current["grp_station"] and local
                                and not _same(local, current["grp_station"])):
                            issues.append("Local Station differs from banner")
                        row["check"] = "yes" if issues else ""
                        row["check_detail"] = "; ".join(issues)
                        row["review_station"] = review or row.get("review_station", "")

                        # Region matching (Sec 19 only, see load_region_map): the
                        # Local Station banner's region code is appended to the
                        # banner value; the mapped name fills Investigating Member,
                        # which a Sec 19 incident is expected to leave blank. If it
                        # wasn't blank, append instead of overwrite and flag the cell
                        # (see write_xlsx) rather than silently discard what was read.
                        if region_map and row.get("sec19") == "Yes":
                            station = current["grp_station"]
                            match = region_map.get(_norm(station).casefold())
                            if match:
                                code, member_name = match
                                row["grp_station"] = f"{station} {code}"
                                row["grp_station_n"] = ""  # split off the banner's count
                                existing = _norm(row.get("investigating_member", ""))
                                if not existing or _is_no_data(existing):
                                    row["investigating_member"] = member_name
                                else:
                                    row["investigating_member"] = f"{existing} / {member_name}"
                                    row["investigating_flag"] = "yes"
                            elif station:
                                print(f"warning: {pdf_path.name} p{i}: Sec 19 "
                                      f"incident's Local Station {station!r} not "
                                      "found in region map — leaving unchanged",
                                      file=sys.stderr)

                        rows.append(row)
                    except Exception as exc:
                        print(f"warning: failed on {pdf_path.name} page {i}: {exc}",
                              file=sys.stderr)
        except Exception as exc:
            print(f"warning: could not open {pdf_path}: {exc}", file=sys.stderr)
    if skipped:
        print(f"Note: skipped {len(skipped)} continuation/non-incident page(s). "
              "See the 'Skipped Pages' sheet to verify them.")
    return rows, skipped


# --------------------------------------------------------------------------- #
#  Writers                                                                     #
# --------------------------------------------------------------------------- #
def _columns() -> list[tuple[str, str]]:
    # "Check" sits right after Page: a single discrepancy flag (per-incident value vs
    # its banner) with no detail — the user follows the Page link to inspect.
    return PROVENANCE + [("check", "Check")] + GROUP + [(f.key, f.header) for f in FIELDS]


def _as_number(value: str):
    """If `value` is a plain number (allowing thousands separators and a €), return
    it as int/float so the cell sorts and right-aligns as money; else None."""
    if not isinstance(value, str):
        return value if isinstance(value, (int, float)) else None
    s = value.replace("€", "").replace(",", "").strip()
    if re.fullmatch(r"-?\d+", s):
        return int(s)
    if re.fullmatch(r"-?\d*\.\d+", s):
        return float(s)
    return None


# Values that mean "the report recorded no data here" (as opposed to a cell the
# parser simply couldn't read, which comes back empty). The report writes these in
# full sentences, e.g. "No investigating officer details found" / "None recorded";
# treating them as no-data lets the writer grey them out so they read at a glance.
_NO_DATA_EXACT = {"n/a", "na", "none", "none recorded", "not recorded",
                  "no record", "no records", "no details"}


def _is_no_data(value: str) -> bool:
    v = _norm(value).lower()
    if not v:
        return False
    if v in _NO_DATA_EXACT:
        return True
    # generic "No … found" / "No … recorded" sentinels the report uses.
    return bool(re.fullmatch(r"no .*(found|recorded)", v))


def _grp_key(r: dict) -> tuple[str, str, str]:
    """The nested group path (why, type, station) of a row — drives the dividers."""
    return tuple(_norm(str(r.get(k, ""))) for k, _l, _h in BANNER_LEVELS)  # type: ignore[return-value]


def _page_link(pdf_path: str, xlsx_path: Path, page) -> str:
    """Hyperlink target that opens `pdf_path` at `page` via the PDF open-parameter
    fragment (`#page=N`). Made relative to the xlsx when possible (so the pair stays
    portable if moved together), else an absolute file URI. Whether the jump lands on
    the right page depends on the PDF viewer honouring #page= (browsers/Acrobat do)."""
    import os
    pdf_abs = os.path.abspath(pdf_path)
    xlsx_dir = os.path.dirname(os.path.abspath(str(xlsx_path)))
    try:
        target = os.path.relpath(pdf_abs, xlsx_dir).replace(os.sep, "/")
    except ValueError:  # e.g. a different drive on Windows
        target = "file:///" + pdf_abs.replace(os.sep, "/").lstrip("/")
    return f"{target}#page={page}"


def write_xlsx(rows: list[dict], path: Path,
               skipped: Optional[list[dict]] = None) -> None:
    """Write a styled, at-a-glance workbook.

    The "Incidents" sheet lists every incident under nested slate dividers — the
    report's three banner levels, Why on Report? > Incident Type > Local Station,
    shaded dark->light by level — with those levels (and Dept) also filled down onto
    every row, so each row stays self-describing. (If no banners are detected it
    falls back to a flat, filterable table.)

    Styling: zebra row banding, the binary Sec 19 / Supp. Role columns shown as
    coloured ✓/✗ (a distinct palette per column), amber for empty cells vs grey for
    recorded "no data", light grid borders and € money formatting. A "Legend" sheet
    documents the colours.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.comments import Comment

    # Columns hidden from the visible grid (CSV/JSON keep the full set), because each
    # duplicates information already on the sheet — so they would only eat print width:
    #   source_file  - identical on every row; shown once in the footer + Legend below
    #                  (and each Page cell still hyperlinks to its own source PDF).
    #   grp_type     - the banner Incident Type; always equals the per-row Incident
    #                  Type field, and still shown as the L2 group divider.
    #   grp_station  - the banner Local Station; the per-row Review Station field is a
    #                  superset of it (station + review station), and it is still shown
    #                  as the L3 group divider.
    #   incident_type - always equals the grp_type banner it falls under (verified at
    #                  parse time; the Check flag fires on any mismatch), shown as the
    #                  L2 group divider.
    #   grp_why      - the Why on Report? banner, shown as the L1 group divider.
    #   grp_dept     - the dept code, constant within a station group and shown in the
    #                  L3 group divider ("Dept …").
    HIDDEN = {"source_file", "grp_type", "grp_station", "incident_type",
              "grp_why", "grp_dept", "check_detail"}
    # The Check column only earns its place when something is actually flagged;
    # with no discrepancies it would just be an empty column, so drop it.
    if not any(_norm(str(r.get("check", ""))).lower() == "yes" for r in rows):
        HIDDEN = HIDDEN | {"check"}
    cols = [c for c in _columns() if c[0] not in HIDDEN]
    ncols = len(cols)
    skipped = skipped or []
    src_files = sorted({Path(str(r["source_file"])).name
                        for r in rows if r.get("source_file")})

    # ---- palette: each kind of cell gets a colour you can read at a glance -- #
    C_HEADER, C_HEADER_TXT = "1F3864", "FFFFFF"   # deep-navy header (anchors the top)
    C_HACC = "8EA9DB"                              # default header underline accent
    C_BAND = "EEF3FA"                              # zebra shading on alt. rows
    C_BLANK = "FFF2CC"                              # amber  -> empty / not extracted
    C_NA_FILL, C_NA_TXT = "E7E6E6", "7F7F7F"       # grey   -> report recorded no data
    C_CHK_FILL, C_CHK_TXT = "FFC7CE", "9C0006"     # red    -> Check: differs from banner
    C_DIV_TXT = "FFFFFF"                            # white text on the divider bars
    C_GRID = "D9D9D9"                              # light cell borders

    thin = Side(style="thin", color=C_GRID)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_border = Border(left=thin, right=thin, top=thin,
                         bottom=Side(style="medium", color=C_HACC))

    band_fill  = PatternFill("solid", fgColor=C_BAND)
    blank_fill = PatternFill("solid", fgColor=C_BLANK)
    na_fill    = PatternFill("solid", fgColor=C_NA_FILL)
    chk_fill   = PatternFill("solid", fgColor=C_CHK_FILL)
    head_fill  = PatternFill("solid", fgColor=C_HEADER)
    C_KEY = "FFE699"                                # gold lane -> Incident No. (key field)
    key_fill = PatternFill("solid", fgColor=C_KEY)
    # Each banner level gets a DISTINCT background hue (not just a lighter slate) so the
    # three levels read as clearly different bands; indent/size reinforce the nesting.
    DIV_SHADE = {"grp_why": "33415C",     # L1 indigo
                 "grp_type": "2E5E4E",    # L2 teal-green
                 "grp_station": "5C3F66"} # L3 plum
    DIV_INDENT = {"grp_why": 0, "grp_type": 1, "grp_station": 2}
    # Banners read as headings: graduated size + row height by level (outer biggest),
    # all clearly larger/taller than the 11pt data rows so they stand out at a glance.
    DIV_SIZE   = {"grp_why": 15, "grp_type": 13, "grp_station": 12}
    DIV_HEIGHT = {"grp_why": 32, "grp_type": 27, "grp_station": 24}
    # The value in each banner is highlighted in a bright colour matched to its level,
    # contrasting against that level's background.
    BANNER_VAL = {"grp_why": "FFCF5C", "grp_type": "A8F0C8", "grp_station": "F0B6E8"}

    base_font = Font(name="Calibri", size=11)
    key_font  = Font(name="Calibri", size=13, bold=True)   # Incident No. — most scanned
    na_font   = Font(name="Calibri", size=11, italic=True, color=C_NA_TXT)
    chk_font  = Font(name="Calibri", size=11, bold=True, color=C_CHK_TXT)
    head_font = Font(bold=True, color=C_HEADER_TXT, size=12)
    div_font  = Font(bold=True, color=C_DIV_TXT, size=11)   # Legend swatches
    div_fonts = {k: Font(name="Calibri", bold=True, color=C_DIV_TXT, size=s)
                 for k, s in DIV_SIZE.items()}
    link_font = Font(name="Calibri", size=11, color="0563C1", underline="single")

    # Binary-state colours — ONLY the noteworthy "yes" state carries colour (a distinct
    # hue per column so Sec 19 and Supp. Role read differently); the routine "no" recedes
    # in grey. {key: {True: (glyph, fill, font), False: (...)}}.
    def _swatch(bg, fg):
        return (PatternFill("solid", fgColor=bg),
                Font(name="Calibri", size=11, bold=True, color=fg))
    _grey = _swatch(C_NA_FILL, C_NA_TXT)
    INDICATOR_STYLE = {
        "sec19": {True:  ("✓", *_swatch("C6EFCE", "006100")),   # green = yes (noteworthy)
                  False: ("✗", *_grey)},                        # grey  = no
        "supp_role_info": {True:  ("✓", *_swatch("BDD7EE", "1F4E79")),  # blue = yes (noteworthy)
                           False: ("✗", *_grey)},                       # grey = no
    }
    # A heavy white rule on the top+bottom ONLY sets each banner apart from the rows
    # around it (and from stacked banner levels). No left/right edges, so the colour
    # runs solid across the whole row instead of being chopped up by cell borders.
    div_edge = Side(style="medium", color="FFFFFF")
    div_border = Border(top=div_edge, bottom=div_edge)

    # Per-column behaviour. Uniform rules so the sheet reads consistently:
    #   * vertical: EVERY data cell is top-aligned (see put_data_row).
    #   * horizontal: RIGHT = money, CENTER = short fixed-shape IDs/codes, else LEFT.
    #   * INDICATOR = binary yes/no fields shown as a tiny centred ✓/✗ (coloured by
    #     state, per INDICATOR_STYLE) in a minimal-width column — no words.
    #   * WIDE = columns shown IN FULL on a SINGLE line (no wrapping); their column is
    #     widened to fit the longest value so nothing is cut off.
    #   * Nothing wraps; any other long value just overflows/clips.
    CENTER = {"page", "incident_no", "case_no", "grp_dept", "reference_no"}
    RIGHT  = {"stolen_eur"}
    INDICATOR = {"sec19", "supp_role_info"}
    WRAP   = set()    # wrapping disabled everywhere
    WIDE   = {"investigating_member", "nominated_supervisor"}
    EMPHASIS = {"incident_no"}   # the most-scanned field: gold lane, bold, larger
    # Each special column's header is underlined in that column's own colour, so the
    # header row doubles as a key for the colour-coding below it; the rest get C_HACC.
    HEADER_ACCENT = {"incident_no": C_KEY, "check": C_CHK_TXT, "sec19": "70AD47",
                     "supp_role_info": "2E75B6", "stolen_eur": "548235"}

    def put_header(ws) -> None:
        ws.append([h for _, h in cols])
        for ci, (key, _) in enumerate(cols, start=1):
            c = ws.cell(row=1, column=ci)
            c.font, c.fill = head_font, head_fill
            accent = HEADER_ACCENT.get(key, C_HACC)
            c.border = Border(left=thin, right=thin, top=thin,
                              bottom=Side(style="thick", color=accent))
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        ws.row_dimensions[1].height = 34

    def _indicator_on(key: str, raw: str) -> bool:
        # Sec 19: a literal "Yes". Supp. Role: any real content means review-worthy
        # supplementary role info is present (None / no-data / blank => no).
        if key == "sec19":
            return raw.strip().lower() == "yes"
        return bool(raw.strip()) and not _is_no_data(raw)

    def put_data_row(ws, rr: int, r: dict, banded: bool) -> None:
        for ci, (key, _) in enumerate(cols, start=1):
            raw_val = r.get(key, "")
            raw = _norm(str(raw_val))

            # Check flag: a red ⚠ when the row's value disagrees with its banner,
            # blank otherwise (no detail — follow the Page link to inspect the source).
            if key == "check":
                on = raw.strip().lower() == "yes"
                cell = ws.cell(row=rr, column=ci, value="⚠" if on else "")
                cell.border = cell_border
                cell.alignment = Alignment(vertical="top", horizontal="center")
                if on:
                    cell.fill, cell.font = chk_fill, chk_font
                    # On-screen hover detail (does NOT print — see Discrepancies sheet).
                    detail = _norm(str(r.get("check_detail", ""))) or "Differs from banner"
                    cm = Comment(detail, "parser"); cm.width, cm.height = 260, 80
                    cell.comment = cm
                else:
                    cell.font = base_font
                    if banded:
                        cell.fill = band_fill
                continue

            # Binary columns: a centred ✓ (yes) / ✗ (no), each state in its own colour
            # (distinct palette per column — see INDICATOR_STYLE), so the zebra band
            # is irrelevant here.
            if key in INDICATOR:
                glyph, fill, font = INDICATOR_STYLE[key][_indicator_on(key, raw)]
                cell = ws.cell(row=rr, column=ci, value=glyph)
                cell.border = cell_border
                cell.alignment = Alignment(vertical="top", horizontal="center")
                cell.fill, cell.font = fill, font
                continue

            value = raw_val
            if key in RIGHT:
                num = _as_number(raw_val)
                value = num if num is not None else raw_val
            cell = ws.cell(row=rr, column=ci, value=value)
            cell.border = cell_border
            cell.font = base_font

            # One alignment rule for all cells: top-aligned vertically; horizontal by
            # data type; wrap only for the show-in-full columns.
            horizontal = ("right" if key in RIGHT else
                          "center" if key in CENTER else "left")
            cell.alignment = Alignment(vertical="top", horizontal=horizontal,
                                       wrap_text=key in WRAP)
            if key in RIGHT and isinstance(value, (int, float)):
                cell.number_format = "€#,##0.00"

            # A Sec 19 incident normally leaves Investigating Member blank; region
            # matching (see parse_files) only appends its mapped name onto a
            # non-blank value, and flags the cell here — same red as a Check
            # discrepancy — so the unexpected pre-existing content stays visible.
            if key == "investigating_member" and \
                    _norm(str(r.get("investigating_flag", ""))) == "yes":
                cell.fill, cell.font = chk_fill, chk_font
            # The key field (Incident No.) gets a persistent gold lane + bold/larger
            # text so the eye lands on it; an empty one still flags amber.
            elif key in EMPHASIS:
                cell.font = key_font
                cell.fill = blank_fill if raw == "" else key_fill
            # Colour, most specific first: empty cells (amber, possible gap), then
            # recorded "no data" (grey), then the zebra band.
            elif raw == "":
                cell.fill = blank_fill
            elif _is_no_data(raw):
                cell.fill, cell.font = na_fill, na_font
            elif banded:
                cell.fill = band_fill

            # Link the Page number to that page of the source PDF (#page=N).
            if key == "page" and raw and r.get("source_path"):
                cell.hyperlink = _page_link(r["source_path"], path, raw)
                cell.font = link_font

    def put_divider(ws, rr: int, level: str, value: str, dept: str,
                    count: int, reported: str) -> None:
        # A slate bar across the row, shaded and indented by banner level; the label
        # lives in column A and overflows rightward over the (empty) cells. No merged
        # cells, so nothing here can interfere with the flat sheet's filter/sort.
        label = next(l for k, l, _h in BANNER_LEVELS if k == level)
        # Trailing bits after the value: dept (station level only) + the incident count.
        tail = []
        if dept and level == "grp_station":
            tail.append(f"Dept {dept}")
        n = f"{count} incident" + ("" if count == 1 else "s")
        if reported and reported.isdigit() and reported != str(count):
            n += f" (banner: {reported})"
        tail.append(n)
        fill = PatternFill("solid", fgColor=DIV_SHADE[level])
        for ci in range(1, ncols + 1):
            c = ws.cell(row=rr, column=ci)
            c.fill, c.border = fill, div_border
            if ci == 1:
                # Indent signals nesting depth. The whole bar is drawn with one
                # cell-level font (white, bold, sized per level) rather than per-run
                # rich text: inline rich text does NOT round-trip reliably through
                # openpyxl, and LibreOffice (the print path) drops per-run
                # formatting — both leave the bar looking like a plain data row.
                indent = "   " * DIV_INDENT[level]
                c.value = (indent + label + " " + (value or "(none)")
                           + "   ·   " + "   ·   ".join(tail))
                c.font = div_fonts[level]
                c.alignment = Alignment(vertical="center", horizontal="left")
        ws.row_dimensions[rr].height = DIV_HEIGHT[level]

    def set_widths(ws) -> None:
        for ci, (key, header) in enumerate(cols, start=1):
            letter = get_column_letter(ci)
            if key == "page":
                ws.column_dimensions[letter].width = 7
                continue
            if key in INDICATOR or key == "check":
                ws.column_dimensions[letter].width = 6   # just a ✓/⚠ — keep it tiny
                continue
            # The header wraps (wrap_text + a tall header row), so every column only
            # has to fit its data and the longest single header *word* — not the whole
            # header string. That stops a short-data column being forced wide by a
            # long heading. WIDE columns get a much higher cap so their full value fits
            # on one line; the rest cap at 40 and clip if longer.
            def _disp_len(v):
                # Money cells render as the formatted string (€, thousands, .00), which
                # is longer than the raw number — size to that so Excel never shows ###.
                if key in RIGHT:
                    n = _as_number(v)
                    if n is not None:
                        return len(f"€{n:,.2f}")
                return len(_norm(str(v)))
            lens = [_disp_len(r.get(key, "")) for r in rows]
            head_word = max(len(w) for w in header.split())
            longest = max([head_word] + lens)
            cap = 80 if key in WIDE else 40
            ws.column_dimensions[letter].width = min(max(longest + 2, 8), cap)

    def span(i: int, depth: int) -> int:  # rows sharing this group prefix from row i
        key = _grp_key(rows[i])[:depth]
        j = i
        while j < len(rows) and _grp_key(rows[j])[:depth] == key:
            j += 1
        return j - i

    wb = Workbook()
    ws = wb.active
    ws.title = "Incidents"
    put_header(ws)

    have_groups = any(any(_grp_key(r)) for r in rows)
    if have_groups:
        # Single sheet: every row sits under the nested banner dividers (Why on
        # Report? > Incident Type > Local Station), each row still self-describing
        # via the filled-down group columns. No auto-filter — the dividers are the
        # organisation here (re-run for a fresh sheet if you need to re-sort).
        rr, band = 1, 0
        prev = ("", "", "")
        for i, r in enumerate(rows):
            cur = _grp_key(r)
            # Re-emit dividers from the outermost level that changed down to station.
            start = next((d for d in range(3) if cur[d] != prev[d]), 3)
            for depth in range(start, 3):
                if not cur[depth]:           # level not present (e.g. before 1st banner)
                    continue
                rr += 1
                key = BANNER_LEVELS[depth][0]
                dept = _norm(str(r.get("grp_dept", ""))) if key == "grp_station" else ""
                reported = _norm(str(r.get(key + "_n", "")))
                put_divider(ws, rr, key, cur[depth], dept, span(i, depth + 1), reported)
                band = 0                     # restart zebra under each fresh group
            rr += 1
            put_data_row(ws, rr, r, band % 2 == 1)
            band += 1
            prev = cur
        ws.freeze_panes = "A2"        # header only (divider text overflows from col A)
    else:
        # No banners detected — fall back to a flat, filterable/sortable table.
        for di, r in enumerate(rows):
            put_data_row(ws, di + 2, r, di % 2 == 1)
        ws.freeze_panes = "B2"        # keep the Page column visible while scrolling
        ws.auto_filter.ref = ws.dimensions
    set_widths(ws)

    # ---- Print as an A4 list: fit every column onto one page wide ---- #
    # File > Print (or Export as PDF) in Excel/LibreOffice then fits the full width
    # to A4 landscape, flowing onto as many pages tall as needed, with the header
    # repeated on each page. Fewer/narrower columns => the fit-to-width scale is
    # gentler => larger print, so the Source File column lives in the footer instead.
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9            # A4
    ws.page_setup.fitToWidth = 1           # all columns -> 1 page wide
    ws.page_setup.fitToHeight = 0          # as many pages tall as needed
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:1"            # repeat the header row on every page
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4,
                                  header=0.2, footer=0.2)
    # Source file(s) once in the footer (left), page numbers on the right.
    if src_files:
        ws.oddFooter.left.text = "Source: " + ", ".join(src_files)
        ws.evenFooter.left.text = ws.oddFooter.left.text
    ws.oddFooter.right.text = "Page &P of &N"
    ws.evenFooter.right.text = ws.oddFooter.right.text

    # ---- Legend sheet: what each colour means ---- #
    leg = wb.create_sheet("Legend")
    leg["A1"] = "Colour key"
    leg["A1"].font = Font(bold=True, size=12)
    shade = {k: PatternFill("solid", fgColor=v) for k, v in DIV_SHADE.items()}
    bval_font = {k: Font(bold=True, color=v) for k, v in BANNER_VAL.items()}
    s19 = INDICATOR_STYLE["sec19"]; sup = INDICATOR_STYLE["supp_role_info"]
    legend_rows = [
        ("✓",    s19[True][1],  s19[True][2],  "Sec 19 = yes (referenced in the narrative)"),
        ("✗",    s19[False][1], s19[False][2], "Sec 19 = no"),
        ("✓",    sup[True][1],  sup[True][2],  "Supp. Role = yes (supplementary role information present)"),
        ("✗",    sup[False][1], sup[False][2], "Supp. Role = no (none recorded)"),
        ("⚠",    chk_fill,  chk_font,  "Check — a value disagrees with the banner it falls under; open the Page link to inspect the source"),
        ("",     blank_fill, base_font, "Empty cell — value missing or not extracted"),
        ("None", na_fill,   na_font,   "Report recorded no data (N/A, None, \"No … recorded/found\")"),
        ("",     band_fill, base_font, "Alternating row shading — easier line tracking"),
        ("L1",  shade["grp_why"],     div_fonts["grp_why"],     "Group divider L1 (outer, largest) — Why on Report?"),
        ("L2",  shade["grp_type"],    div_fonts["grp_type"],    "Group divider L2 — Incident Type"),
        ("L3",  shade["grp_station"], div_fonts["grp_station"], "Group divider L3 (inner) — Local Station (with dept + count)"),
        ("Aa",  shade["grp_why"],     bval_font["grp_why"],     "Banner value colour — Why on Report? (level 1)"),
        ("Aa",  shade["grp_type"],    bval_font["grp_type"],    "Banner value colour — Incident Type (level 2)"),
        ("Aa",  shade["grp_station"], bval_font["grp_station"], "Banner value colour — Local Station (level 3)"),
    ]
    for i, (sample, fill, font, meaning) in enumerate(legend_rows, start=3):
        sc = leg.cell(row=i, column=1, value=sample)
        sc.fill, sc.font, sc.border = fill, font, cell_border
        sc.alignment = Alignment(horizontal="center", vertical="center")
        mc = leg.cell(row=i, column=2, value=meaning)
        mc.alignment = Alignment(vertical="center")
    leg.column_dimensions["A"].width = 10
    leg.column_dimensions["B"].width = 66
    leg.sheet_view.showGridLines = False

    # Source file(s) on the Legend too, so the provenance is visible on screen
    # (the print footer only shows when printing/exporting to PDF).
    if src_files:
        sr = len(legend_rows) + 4
        leg.cell(row=sr, column=1, value="Source file(s):").font = Font(bold=True)
        leg.cell(row=sr, column=2, value=", ".join(src_files)).alignment = \
            Alignment(vertical="center", wrap_text=True)

    # ---- Summary sheet (first): title + at-a-glance breakdown, prints first ---- #
    import datetime
    from collections import Counter
    flagged = [r for r in rows if _norm(str(r.get("check", ""))).lower() == "yes"]
    sec19_yes = sum(1 for r in rows if _norm(str(r.get("sec19", ""))).lower() == "yes")
    supp_yes = sum(1 for r in rows if _indicator_on("supp_role_info",
                                                    _norm(str(r.get("supp_role_info", "")))))
    title_font = Font(bold=True, size=18, color="1F3864")
    sect_font  = Font(bold=True, size=12, color="FFFFFF")
    sect_fill  = PatternFill("solid", fgColor=C_HEADER)
    label_font = Font(bold=True)
    big_num    = Font(bold=True, size=14)

    summ = wb.create_sheet("Summary", 0)
    summ.sheet_view.showGridLines = False
    summ["A1"] = "Incident Report — Condensed"
    summ["A1"].font = title_font
    summ["A2"] = f"Generated {datetime.date.today().isoformat()}"
    summ["A2"].font = Font(italic=True, color="7F7F7F")
    if src_files:
        summ["A3"] = "Source: " + ", ".join(src_files)
        summ["A3"].font = Font(color="7F7F7F")

    def _section(r, text):
        c = summ.cell(row=r, column=1, value=text)
        c.font, c.fill = sect_font, sect_fill
        summ.cell(row=r, column=2).fill = sect_fill
        return r + 1

    def _stat(r, label, value):
        summ.cell(row=r, column=1, value=label).font = label_font
        summ.cell(row=r, column=2, value=value).font = big_num
        return r + 1

    def _breakdown(r, title, counts):
        r = _section(r, title)
        for name, n in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])):
            summ.cell(row=r, column=1, value=name or "(none)")
            summ.cell(row=r, column=2, value=n).alignment = Alignment(horizontal="right")
            r += 1
        return r + 1

    r = 5
    r = _section(r, "Totals")
    r = _stat(r, "Incidents", len(rows))
    r = _stat(r, "Flagged for check ⚠", len(flagged))
    r = _stat(r, "Sec 19 referenced", sec19_yes)
    r = _stat(r, "Supplementary role info", supp_yes)
    r = _stat(r, "Pages skipped (see sheet)", len(skipped))
    r += 1
    r = _breakdown(r, "By Why on Report?",
                   Counter(_norm(str(x.get("grp_why", ""))) for x in rows))
    r = _breakdown(r, "By Incident Type",
                   Counter(_norm(str(x.get("grp_type", ""))) for x in rows))
    summ.column_dimensions["A"].width = 40
    summ.column_dimensions["B"].width = 14
    summ.page_setup.paperSize = 9          # A4
    summ.page_setup.fitToWidth = 1
    summ.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    # ---- Discrepancies sheet (last): the printable detail behind each ⚠ flag ---- #
    disc = wb.create_sheet("Discrepancies")
    disc.sheet_view.showGridLines = False
    disc["A1"] = "Discrepancies — values that differ from the banner above them"
    disc["A1"].font = Font(bold=True, size=14, color=C_CHK_TXT)
    disc["A2"] = "Each row below is flagged ⚠ on the Incidents sheet. Click the Page to open the source PDF."
    disc["A2"].font = Font(italic=True, color="7F7F7F")
    if flagged:
        dh = ["Page", "Incident No.", "Case No.", "Issue"]
        for ci, h in enumerate(dh, start=1):
            c = disc.cell(row=4, column=ci, value=h)
            c.font, c.fill, c.border = head_font, head_fill, head_border
            c.alignment = Alignment(horizontal="center", vertical="center")
        dr = 5
        for r_ in flagged:
            pg = _norm(str(r_.get("page", "")))
            pc = disc.cell(row=dr, column=1, value=_as_number(pg) or pg)
            pc.alignment = Alignment(horizontal="center")
            if pg and r_.get("source_path"):
                pc.hyperlink = _page_link(r_["source_path"], path, pg)
                pc.font = link_font
            disc.cell(row=dr, column=2, value=r_.get("incident_no", "")).alignment = \
                Alignment(horizontal="center")
            disc.cell(row=dr, column=3, value=r_.get("case_no", "")).alignment = \
                Alignment(horizontal="center")
            disc.cell(row=dr, column=4, value=_norm(str(r_.get("check_detail", ""))))
            for ci in range(1, 5):
                disc.cell(row=dr, column=ci).border = cell_border
            dr += 1
        disc.print_title_rows = "4:4"
    else:
        disc["A4"] = "No discrepancies found."
        disc["A4"].font = Font(bold=True, color="006100")
    for col, w in (("A", 8), ("B", 14), ("C", 14), ("D", 52)):
        disc.column_dimensions[col].width = w
    disc.page_setup.paperSize = 9          # A4
    disc.page_setup.orientation = "landscape"
    disc.page_setup.fitToWidth = 1
    disc.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    # ---- Skipped Pages sheet (last): pages NOT extracted, so they can be verified -- #
    skp = wb.create_sheet("Skipped Pages")
    skp.sheet_view.showGridLines = False
    skp["A1"] = "Skipped pages — not extracted (treated as continuation / non-incident)"
    skp["A1"].font = Font(bold=True, size=14, color="1F3864")
    skp["A2"] = ("Click each Page to open it in the source PDF and confirm it holds no "
                 "incident that should have been captured.")
    skp["A2"].font = Font(italic=True, color="7F7F7F")
    if skipped:
        for ci, h in enumerate(["Page", "Source File"], start=1):
            c = skp.cell(row=4, column=ci, value=h)
            c.font, c.fill, c.border = head_font, head_fill, head_border
            c.alignment = Alignment(horizontal="center", vertical="center")
        sdr = 5
        for s in skipped:
            pg = _norm(str(s.get("page", "")))
            pc = skp.cell(row=sdr, column=1, value=_as_number(pg) or pg)
            pc.alignment = Alignment(horizontal="center")
            if pg and s.get("source_path"):
                pc.hyperlink = _page_link(s["source_path"], path, pg)
                pc.font = link_font
            skp.cell(row=sdr, column=2, value=s.get("source_file", ""))
            for ci in range(1, 3):
                skp.cell(row=sdr, column=ci).border = cell_border
            sdr += 1
        skp.print_title_rows = "4:4"
    else:
        skp["A4"] = "No pages were skipped — every page was extracted."
        skp["A4"].font = Font(bold=True, color="006100")
    skp.column_dimensions["A"].width = 8
    skp.column_dimensions["B"].width = 40
    skp.page_setup.paperSize = 9           # A4
    skp.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    wb.active = 0   # open on the Summary
    wb.save(path)


def write_csv(rows: list[dict], path: Path,
              skipped: Optional[list[dict]] = None) -> None:
    cols = _columns()
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow([h for _, h in cols])
        for r in rows:
            w.writerow([r.get(k, "") for k, _ in cols])


def write_json(rows: list[dict], path: Path,
               skipped: Optional[list[dict]] = None) -> None:
    # Project to the defined columns so internal stash keys (e.g. the banner-stated
    # counts grp_*_n) stay out of the export, matching the CSV/xlsx contract.
    keys = [k for k, _ in _columns()]
    out = [{k: r.get(k, "") for k in keys} for r in rows]
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(out, fh, ensure_ascii=False, indent=2)


WRITERS = {"xlsx": write_xlsx, "csv": write_csv, "json": write_json}
DEFAULT_OUT = {"xlsx": "incidents.xlsx", "csv": "incidents.csv", "json": "incidents.json"}


# --------------------------------------------------------------------------- #
#  Self-test against the bundled template                                      #
# --------------------------------------------------------------------------- #
def self_test() -> int:
    template = Path(__file__).with_name("template.pdf")
    if not template.exists():
        print(f"self-test: {template} not found", file=sys.stderr)
        return 2
    rows, _ = parse_files([template])
    if not rows:
        print("self-test: no rows extracted", file=sys.stderr)
        return 1
    row = rows[0]
    # In the template every *** target cell literally contains the marker.
    expect_marker = {"incident_no", "incident_type", "investigating_member",
                     "nominated_supervisor", "stolen_eur"}
    # Cells that are not "***" in the template but whose value is known. The
    # Narrative cell is "***", so it references no Sec 19 and holds no F…x… number;
    # the "Station (Review Station)" column reads "place (place)" and loc_review_station
    # keeps only the bracketed Review Station -> "place", confirming both that the
    # locator reads the right column and that the bracket extraction works.
    expect_value = {"case_no": "None", "sec19": "No", "reference_no": "N/A",
                    "review_station": "place"}
    ok = True
    print("self-test (template values):")
    for f in FIELDS:
        val = row[f.key]
        if f.key in expect_marker:
            good = val == "***"
        elif f.key in expect_value:
            good = val == expect_value[f.key]
        else:  # supp_role_info: red cell text -> '*** (needed if cell is red)'
            good = "***" in val
        ok = ok and good
        print(f"  [{'PASS' if good else 'FAIL'}] {f.key:22s} = {val!r}")
    print("RESULT:", "PASS" if ok else "FAIL")
    return 0 if ok else 1


# --------------------------------------------------------------------------- #
#  Privacy-safe inspection (for sharing structure without data)                #
# --------------------------------------------------------------------------- #
def _mask(value: str) -> str:
    """Describe a cell value by SHAPE only — length + character classes, never
    the text. Lets an inspection report be shared without leaking any data."""
    if not value:
        return "<empty>"
    cls = []
    if any(c.isalpha() for c in value):
        cls.append("letters")
    if any(c.isdigit() for c in value):
        cls.append("digits")
    if any(c in "*#€" for c in value):
        cls.append("marks")
    if any(not c.isalnum() and not c.isspace() and c not in "*#€" for c in value):
        cls.append("punct")
    return f"<{len(value)} chars: {'+'.join(cls) or 'other'}>"


def inspect_files(paths: list[Path]) -> int:
    """Print a privacy-safe structural report: per page, its classification, which
    KNOWN template section labels are present, and the SHAPE of each field value
    (masked). Nothing from the page body is printed verbatim, so the output is
    safe to share for tuning the parser on layouts I can't see."""
    pages = 0
    for pdf_path in paths:
        try:
            pdf = pdfplumber.open(pdf_path)
        except Exception as exc:
            print(f"warning: could not open {pdf_path}: {exc}", file=sys.stderr)
            continue
        with pdf:
            print(f"\n===== {pdf_path.name} =====")
            for i, page in enumerate(pdf.pages, start=1):
                pages += 1
                pv = PageView(page)
                matched = [a for a in _SECTION_ANCHORS if pv.find(a) is not None]
                flag = is_incident_page(pv)
                print(f"\nPAGE {i}  {pv.width:.0f}x{pv.height:.0f}pt  "
                      f"words={len(pv.words)}  ->  "
                      f"{'INCIDENT' if flag else 'NON-INCIDENT / continuation'} "
                      f"(known sections {len(matched)}/{len(_SECTION_ANCHORS)})")
                print("  sections present:",
                      ", ".join(matched) if matched
                      else "none of the known template labels")
                for f in FIELDS:
                    try:
                        val = f.locate(pv)
                    except Exception:
                        val = ""
                    print(f"    {f.key:22s} {_mask(val)}")
    if not pages:
        print("error: no pages found", file=sys.stderr)
        return 1
    print(f"\nInspected {pages} page(s). Values are masked (shape only) — this "
          "report contains no page data and is safe to share.\nFor any page shown "
          "as NON-INCIDENT, the section/column headings printed by your viewer are "
          "field labels (not data); tell me those and I can add support.")
    return 0


# --------------------------------------------------------------------------- #
#  Text-layer diagnosis (why do fields come back empty?)                       #
# --------------------------------------------------------------------------- #
# Single-word tokens that appear on a normal incident page. How many of these
# are extractable is a quick gauge of text-layer health (no page data revealed).
_PROBE_WORDS = (
    "Incident", "Type", "Occurred", "Reported", "Scene", "Narrative", "Quality",
    "Control", "Tests", "Role", "Details", "Created", "Reporting", "Investigating",
    "Nominated", "Station", "Outcome", "Case", "Local", "Supplementary",
)


def diagnose_files(paths: list[Path]) -> int:
    """Privacy-safe text-layer health check. The usual reason a field is empty
    although the page clearly shows text is that the PDF has no real text layer
    (a scan/screenshot) or a broken font encoding. This reports only counts and
    font/image stats — never page content — and gives a per-page verdict."""
    for pdf_path in paths:
        try:
            pdf = pdfplumber.open(pdf_path)
        except Exception as exc:
            print(f"warning: could not open {pdf_path}: {exc}", file=sys.stderr)
            continue
        with pdf:
            print(f"\n===== {pdf_path.name} =====")
            for i, page in enumerate(pdf.pages, start=1):
                chars = page.chars or []
                words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
                fonts = {str(c.get("fontname", "?")).split("+")[-1] for c in chars}
                cid = sum(1 for c in chars if str(c.get("text", "")).startswith("(cid:"))
                area = (page.width or 1) * (page.height or 1)
                cover = max((abs(im["width"] * im["height"])
                             for im in (page.images or [])), default=0) / area * 100
                wordset = {w["text"] for w in words}
                hits = sum(1 for t in _PROBE_WORDS if t in wordset)
                if len(words) < 40 and cover > 60:
                    verdict = "IMAGE/SCAN — little/no extractable text; needs OCR"
                elif chars and cid > 0.2 * len(chars):
                    verdict = "BROKEN FONT ENCODING — text present but not mapped to Unicode"
                elif hits >= len(_PROBE_WORDS) * 0.6:
                    verdict = "TEXT LAYER OK"
                else:
                    verdict = "SPARSE/PARTIAL text layer (e.g. incomplete OCR)"
                print(f"\nPAGE {i}: words={len(words)} chars={len(chars)} "
                      f"fonts={len(fonts)} broken_glyphs={cid} "
                      f"largest_image={cover:.0f}% of page  "
                      f"template_words={hits}/{len(_PROBE_WORDS)}")
                print(f"  -> {verdict}")
    print("\n(Counts and font/image stats only — no page content is shown, so this "
          "is safe to share. 'TEXT LAYER OK' means fields should extract; the other "
          "verdicts mean the text isn't really in the PDF and must be OCR'd or "
          "re-exported as a text PDF.)")
    return 0


# --------------------------------------------------------------------------- #
#  CLI                                                                         #
# --------------------------------------------------------------------------- #
def main(argv: Optional[list[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        description="Condense structured incident-report PDFs into a "
        "single spreadsheet (extracts the *** cells).")
    ap.add_argument("inputs", nargs="*",
                    help="PDF file(s) and/or folder(s) of PDFs")
    ap.add_argument("-o", "--output", help="output file path")
    ap.add_argument("-f", "--format", choices=list(WRITERS), default="xlsx",
                    help="output format (default: xlsx)")
    ap.add_argument("--debug", action="store_true",
                    help="print every extracted field for each page")
    ap.add_argument("--self-test", action="store_true",
                    help="run the parser against the bundled template.pdf and exit")
    ap.add_argument("--inspect", action="store_true",
                    help="print a privacy-safe structural report (page types and "
                         "field shapes, all values masked) — safe to share for tuning")
    ap.add_argument("--diagnose", action="store_true",
                    help="check whether a PDF's text is actually extractable "
                         "(scan/OCR/encoding health) — safe to share, no content")
    ap.add_argument("--region-map", metavar="FILE",
                    help="Local Station -> region code/name map for Sec 19 "
                         "incidents — plaintext or an --encrypt-region-map output "
                         "(see parse_region_map for the format); auto-detected next "
                         "to this script if present (encrypted preferred over "
                         "plaintext), otherwise the feature is off")
    ap.add_argument("--encrypt-region-map", metavar="FILE", nargs="?",
                    const="region_stations.md",
                    help="encrypt FILE (default: region_stations.md) to FILE.enc "
                         "with a passphrase, then exit — delete the plaintext "
                         "original yourself once the encrypted copy loads correctly")
    args = ap.parse_args(argv)

    if args.self_test:
        return self_test()

    if args.encrypt_region_map:
        src = Path(args.encrypt_region_map)
        if not src.exists():
            ap.error(f"--encrypt-region-map: {src} not found")
        p1 = getpass.getpass(f"Passphrase for {src.name}: ")
        p2 = getpass.getpass("Confirm passphrase: ")
        if not p1:
            print("error: empty passphrase", file=sys.stderr)
            return 1
        if p1 != p2:
            print("error: passphrases did not match", file=sys.stderr)
            return 1
        dest = src.with_name(src.name + ".enc")
        encrypt_region_map(src, dest, p1)
        print(f"Encrypted -> {dest}\n"
              f"Verify it loads (--region-map {dest}), then delete the plaintext "
              f"{src} yourself — it is not removed automatically.")
        return 0

    if not args.inputs:
        ap.error("no inputs given (provide PDF files/folders, or use --self-test)")

    paths = iter_pdf_paths(args.inputs)
    if not paths:
        print("error: no PDF files found in the given inputs", file=sys.stderr)
        return 1

    if args.inspect:
        return inspect_files(paths)

    if args.diagnose:
        return diagnose_files(paths)

    region_map = None
    if args.region_map:
        region_map_path = Path(args.region_map)
        if not region_map_path.exists():
            ap.error(f"--region-map file not found: {region_map_path}")
    else:
        here = Path(__file__).resolve().parent
        enc, plain = here / "region_stations.md.enc", here / "region_stations.md"
        region_map_path = enc if enc.exists() else (plain if plain.exists() else None)

    if region_map_path is not None:
        if region_map_path.suffix == ".enc":
            passphrase = getpass.getpass(f"Passphrase for {region_map_path.name}: ")
            try:
                region_map = decrypt_region_map(region_map_path, passphrase)
            except ValueError as exc:
                print(f"error: {exc}", file=sys.stderr)
                return 1
        else:
            print(f"warning: {region_map_path} is unencrypted plaintext on disk "
                  "— consider --encrypt-region-map", file=sys.stderr)
            region_map = load_region_map(region_map_path)
        print(f"Using region map: {region_map_path} ({len(region_map)} station(s))")

    print(f"Parsing {len(paths)} PDF file(s)…")
    rows, skipped = parse_files(paths, debug=args.debug, region_map=region_map)
    if not rows:
        print("error: nothing extracted", file=sys.stderr)
        return 1

    out = Path(args.output) if args.output else Path(DEFAULT_OUT[args.format])
    WRITERS[args.format](rows, out, skipped=skipped)
    print(f"Done: {len(rows)} incident page(s) from {len(paths)} file(s) "
          f"-> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
