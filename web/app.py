"""Stateless web front-end for the incident-report parser.

Design:
  * /parse  - accept an uploaded report PDF, parse it in a per-request temp dir
              (0700, deleted in a finally), and return the .xlsx bytes. Nothing is
              persisted: the browser holds the only copy of the result.
  * /print  - accept the .xlsx the browser is holding, render it to PDF with
              LibreOffice headless and send it to the network printer over IPP.
              Again a per-request temp dir, deleted afterwards.
  * /        - the single-page UI.
  * /healthz - liveness/readiness for Kubernetes (unauthenticated).

Privacy: this handles sensitive personal data. We never write outside a private
per-request temp dir, never log file *contents*, and strip the parser's internal
`source_path` so no server-side path leaks into the .xlsx hyperlinks.

Security (see SECURITY.md for the full model):
  * Identity is enforced *in front* by an auth proxy (Cloudflare Access). As defence
    in depth this app can ALSO verify the Access JWT itself (set ACCESS_TEAM_DOMAIN
    + ACCESS_AUD) so a request that reaches the pod directly — bypassing the proxy —
    is still rejected. Disabled by default; a warning is logged when off. (H2)
  * /print only accepts a workbook THIS server produced: /parse returns an HMAC over
    the bytes and /print re-checks it before the file touches openpyxl/LibreOffice,
    so no arbitrary document can be fed to LibreOffice. (M1)
  * Concurrency for the CPU/subprocess-heavy work is capped, and child processes run
    in their own group so a timeout kills the whole tree. (M2)
  * Formula/CSV injection in the produced workbook is neutralised in the parser
    (parse_incidents._formula_safe); relevant because /print renders it through
    LibreOffice. (H1)
"""
from __future__ import annotations

import asyncio
import hashlib
import hmac
import logging
import os
import re
import secrets
import shutil
import signal
import subprocess
import sys
import tempfile
import unicodedata
from pathlib import Path

from fastapi import Depends, FastAPI, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, Response

# The parser lives at the repo root, one level above this file's directory.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import parse_incidents as parser  # noqa: E402

# ---- configuration (all via env so the image is config-free) -----------------
MAX_UPLOAD_MB = int(os.environ.get("MAX_UPLOAD_MB", "25"))
MAX_UPLOAD_BYTES = MAX_UPLOAD_MB * 1024 * 1024
# Direct IPP Everywhere endpoint of the network printer, e.g.
# "ipp://<printer-host>/ipp/print" (or "ipps://…" for TLS on the wire — preferred
# for sensitive output). Printing straight to the printer means we don't depend on
# any intermediate CUPS host being powered on.
PRINTER_URI = os.environ.get("PRINTER_URI", "")
IPP_TEMPLATE = Path(__file__).parent / "ipp" / "print-job.test"
SOFFICE_BIN = os.environ.get("SOFFICE_BIN", "soffice")
SOFFICE_TIMEOUT = int(os.environ.get("SOFFICE_TIMEOUT", "120"))
PRINT_TIMEOUT = int(os.environ.get("PRINT_TIMEOUT", "60"))
# Optional Sec 19 region map (see parse_incidents.load_region_map). Not baked into
# the image or this repo — supplied at deploy time as a mounted Secret file, so
# the real station/region names never appear in either public repo or the public
# image. Absent by default: the feature is then simply off.
REGION_MAP_PATH = Path(os.environ.get("REGION_MAP_PATH", "/etc/region-map/region_stations.md"))

# Cap concurrent CPU/subprocess-heavy work (pdfplumber parse, LibreOffice render) so
# a burst of large/complex uploads can't exhaust the pod; requests queue on the
# semaphore. k8s resource limits are the outer bound. (M2)
MAX_CONCURRENCY = int(os.environ.get("MAX_CONCURRENCY", "2"))
_WORK_SEM = asyncio.Semaphore(MAX_CONCURRENCY)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("file-parser-web")

# ---- /print binding key (M1) -------------------------------------------------
# Prefer a stable key (survives restarts / works across replicas); otherwise a
# per-process key — fine for a single replica, but a workbook downloaded before a
# restart then can't be printed (just re-parse).
_key_env = os.environ.get("PRINT_SIGNING_KEY", "").encode()
_EPHEMERAL_PRINT_KEY = not _key_env
_PRINT_KEY = _key_env or secrets.token_bytes(32)
if _EPHEMERAL_PRINT_KEY:
    log.warning("PRINT_SIGNING_KEY not set — using an ephemeral per-process print "
                "token key (ok for a single replica; set it for multi-replica or "
                "print-after-restart).")

# ---- optional Cloudflare Access JWT verification (H2) -------------------------
ACCESS_TEAM_DOMAIN = os.environ.get("ACCESS_TEAM_DOMAIN", "").strip().rstrip("/")
ACCESS_AUD = os.environ.get("ACCESS_AUD", "").strip()
ACCESS_ENABLED = bool(ACCESS_TEAM_DOMAIN and ACCESS_AUD)
_ISSUER = ACCESS_TEAM_DOMAIN if ACCESS_TEAM_DOMAIN.startswith("http") else f"https://{ACCESS_TEAM_DOMAIN}"
_jwk_client = None
if ACCESS_ENABLED:
    # Fail closed: if verification was explicitly requested but the library is
    # missing/misconfigured, refuse to start rather than run without the check.
    from jwt import PyJWKClient  # noqa: E402  (raises at import if PyJWT absent)
    _jwk_client = PyJWKClient(f"{_ISSUER}/cdn-cgi/access/certs")  # caches keys itself
    log.info("Cloudflare Access JWT verification ENABLED (issuer=%s)", _ISSUER)
else:
    log.warning("Cloudflare Access JWT verification DISABLED — relying solely on the "
                "auth proxy + NetworkPolicy in front. Set ACCESS_TEAM_DOMAIN + "
                "ACCESS_AUD to enable in-app verification.")


def _verify_access(request: Request) -> str:
    """Return the authenticated email. When Access verification is enabled, the
    Cloudflare Access JWT is cryptographically verified (signature, audience,
    issuer); a missing/invalid assertion is rejected. When disabled, fall back to
    the (untrusted) header Cloudflare injects — usable for local dev, but see the
    startup warning."""
    if not ACCESS_ENABLED:
        return request.headers.get("Cf-Access-Authenticated-User-Email", "anonymous")
    import jwt
    token = (request.headers.get("Cf-Access-Jwt-Assertion")
             or request.cookies.get("CF_Authorization"))
    if not token:
        raise HTTPException(401, "Missing Cloudflare Access assertion.")
    try:
        key = _jwk_client.get_signing_key_from_jwt(token).key
        claims = jwt.decode(token, key, algorithms=["RS256"],
                            audience=ACCESS_AUD, issuer=_ISSUER)
    except Exception as exc:  # signature / aud / issuer / expiry / fetch failure
        log.warning("Access JWT rejected: %s", exc.__class__.__name__)
        raise HTTPException(401, "Invalid Cloudflare Access assertion.")
    return claims.get("email") or claims.get("common_name") or "authenticated"


app = FastAPI(title="Incident report parser", docs_url=None, redoc_url=None)

_INDEX_HTML = (Path(__file__).parent / "templates" / "index.html").read_text()

# Loaded once at startup, not per-request: a mounted Secret file doesn't change
# without a pod restart anyway (same as the ConfigMap-sourced env vars above).
# Never logs the map's contents, only its size, matching the CLI's own logging.
REGION_MAP: dict[str, tuple[str, str]] | None = None
if REGION_MAP_PATH.exists():
    try:
        REGION_MAP = parser.load_region_map(REGION_MAP_PATH)
        log.info("region map loaded: %d station(s)", len(REGION_MAP))
    except Exception:
        log.exception("failed to load region map from %s — feature disabled", REGION_MAP_PATH)


@app.middleware("http")
async def _security_headers(request: Request, call_next):
    """Set response security headers, incl. a per-request CSP nonce the index page
    stamps into its inline <style>/<script> so no 'unsafe-inline' is needed. (L3)"""
    nonce = secrets.token_urlsafe(16)
    request.state.csp_nonce = nonce
    response = await call_next(request)
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; "
        f"script-src 'nonce-{nonce}'; style-src 'nonce-{nonce}'; "
        "img-src 'self' data:; connect-src 'self'; object-src 'none'; "
        "base-uri 'none'; form-action 'self'; frame-ancestors 'none'")
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "no-referrer"
    response.headers["Cross-Origin-Opener-Policy"] = "same-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    return response


def _private_tmpdir() -> str:
    """A 0700 temp dir for one request's files; caller must rmtree it."""
    return tempfile.mkdtemp(prefix="fp-")


def _safe_download_name(name: str) -> str:
    """A filesystem/header-safe download name derived from the upload, so the
    user-controlled filename can't break the Content-Disposition header (CR/LF,
    embedded quotes) or trip latin-1 encoding (non-ASCII → 500). (L1)"""
    stem = Path(name).stem or "report"
    stem = unicodedata.normalize("NFKD", stem).encode("ascii", "ignore").decode()
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._") or "report"
    return f"{stem[:100]}-incidents.xlsx"


def _run(cmd: list[str], timeout: int, **kwargs):
    """Run a subprocess in its own session/process group so a timeout can kill the
    whole tree — LibreOffice forks soffice.bin, which subprocess.run's timeout would
    orphan. Returns (returncode, stdout, stderr); re-raises TimeoutExpired. (M2)"""
    proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                            text=True, start_new_session=True, **kwargs)
    try:
        out, err = proc.communicate(timeout=timeout)
    except subprocess.TimeoutExpired:
        try:
            os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
        except (ProcessLookupError, PermissionError):
            pass
        proc.communicate()
        raise
    return proc.returncode, out, err


PRINT_SHEET = "Incidents"


def _incidents_only(src: Path) -> Path:
    """Return a copy of the workbook containing only the `Incidents` sheet, so the
    printout omits Summary/Legend/Discrepancies/Skipped Pages. If that sheet isn't
    present (unexpected), fall back to printing the workbook unchanged."""
    from openpyxl import load_workbook

    wb = load_workbook(src)
    if PRINT_SHEET not in wb.sheetnames:
        return src
    for name in [n for n in wb.sheetnames if n != PRINT_SHEET]:
        del wb[name]
    out = src.with_name("print.xlsx")
    wb.save(out)
    return out


async def _save_upload(upload: UploadFile, dest: Path) -> int:
    """Stream an upload to `dest`, enforcing the size cap. Returns bytes written."""
    written = 0
    with dest.open("wb") as fh:
        while chunk := await upload.read(1024 * 1024):
            written += len(chunk)
            if written > MAX_UPLOAD_BYTES:
                raise HTTPException(413, f"File exceeds {MAX_UPLOAD_MB} MB limit.")
            fh.write(chunk)
    return written


@app.get("/", response_class=HTMLResponse)
async def index(request: Request) -> HTMLResponse:
    return HTMLResponse(_INDEX_HTML.replace("__CSP_NONCE__", request.state.csp_nonce))


@app.get("/healthz")
async def healthz() -> JSONResponse:
    return JSONResponse({"status": "ok"})


@app.post("/parse")
async def parse(request: Request, file: UploadFile,
                user: str = Depends(_verify_access)) -> Response:
    name = file.filename or "report.pdf"
    if not name.lower().endswith(".pdf"):
        raise HTTPException(400, "Please upload a .pdf export of the report.")
    work = _private_tmpdir()
    try:
        pdf_path = Path(work) / "input.pdf"
        size = await _save_upload(file, pdf_path)

        async with _WORK_SEM:
            rows, skipped = parser.parse_files([pdf_path], region_map=REGION_MAP)
        if not rows:
            raise HTTPException(
                422,
                "No incidents were extracted. If the PDF looks full, it is probably "
                "a print-to-PDF or scan with no text layer — re-export it from the "
                "report system (or OCR it first).",
            )
        # Strip the parser's internal absolute path so it can't leak into the
        # workbook's Page hyperlinks (those point at this server's temp dir, which
        # the recipient can't open and shouldn't see). write_xlsx skips the
        # hyperlink when source_path is absent.
        for r in rows:
            r.pop("source_path", None)
        for s in skipped:
            s.pop("source_path", None)

        out_path = Path(work) / "incidents.xlsx"
        parser.write_xlsx(rows, out_path, skipped)
        data = out_path.read_bytes()

        # Bind this workbook to /print (M1): the browser must send this token back.
        token = hmac.new(_PRINT_KEY, data, hashlib.sha256).hexdigest()

        log.info("parse user=%s in=%dB rows=%d skipped=%d out=%dB",
                 user, size, len(rows), len(skipped), len(data))

        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{_safe_download_name(name)}"',
                "X-Incident-Rows": str(len(rows)),
                "X-Skipped-Pages": str(len(skipped)),
                "X-Print-Token": token,
            },
        )
    finally:
        shutil.rmtree(work, ignore_errors=True)


@app.post("/print")
async def print_xlsx(request: Request, file: UploadFile, token: str = Form(""),
                     mono: str = Form(""),
                     user: str = Depends(_verify_access)) -> JSONResponse:
    if not PRINTER_URI:
        raise HTTPException(503, "Printing is not configured on this server.")
    # Black & white when requested, else let the printer decide (colour for a
    # colour printer). This maps a browser flag onto a fixed IPP keyword — the
    # value is never taken verbatim from the request, so no injection into ipptool.
    color_mode = "monochrome" if mono.strip().lower() in {"1", "true", "on", "yes"} else "auto"
    work = _private_tmpdir()
    try:
        xlsx_path = Path(work) / "incidents.xlsx"
        await _save_upload(file, xlsx_path)

        # M1: only ever print a workbook this server produced. Verify the HMAC the
        # browser got from /parse BEFORE the file touches openpyxl or LibreOffice,
        # so an attacker (past the auth proxy) can't feed an arbitrary document to
        # LibreOffice. Constant-time compare; reject if absent/mismatched.
        data = xlsx_path.read_bytes()
        expected = hmac.new(_PRINT_KEY, data, hashlib.sha256).hexdigest()
        if not hmac.compare_digest(expected, token or ""):
            log.warning("print rejected (bad/absent token) user=%s", user)
            raise HTTPException(
                403, "This file was not produced by this server, or the session "
                "expired. Please re-parse the PDF and print again.")

        async with _WORK_SEM:
            # Print the "Incidents" sheet only — LibreOffice renders every sheet, so
            # drop the rest from a throwaway copy first. The user's downloaded
            # workbook is untouched; this copy exists only to print.
            xlsx_path = _incidents_only(xlsx_path)

            # Render to PDF. A throwaway LibreOffice profile per call keeps concurrent
            # requests from fighting over one user profile; HOME is the temp dir.
            profile = Path(work) / "lo-profile"
            try:
                rc, _out, err = _run(
                    [SOFFICE_BIN, "--headless", "--norestore",
                     f"-env:UserInstallation=file://{profile}",
                     "--convert-to", "pdf", "--outdir", work, str(xlsx_path)],
                    timeout=SOFFICE_TIMEOUT, env={**os.environ, "HOME": work})
            except subprocess.TimeoutExpired:
                log.error("soffice timed out after %ss", SOFFICE_TIMEOUT)
                raise HTTPException(504, "Rendering the workbook timed out.")
            pdf_path = xlsx_path.with_suffix(".pdf")
            if rc != 0 or not pdf_path.exists():
                log.error("soffice failed: %s", (err or "").strip())
                raise HTTPException(500, "Could not render the workbook for printing.")

            # Submit the PDF straight to the printer's IPP endpoint. ipptool returns
            # non-zero unless the Print-Job's STATUS expectation (successful-ok) holds.
            try:
                prc, pout, perr = _run(
                    ["ipptool", "-tv", "-d", f"color_mode={color_mode}",
                     "-f", str(pdf_path), PRINTER_URI, str(IPP_TEMPLATE)],
                    timeout=PRINT_TIMEOUT)
            except subprocess.TimeoutExpired:
                log.error("ipptool timed out after %ss", PRINT_TIMEOUT)
                raise HTTPException(504, "The printer did not respond in time.")
            if prc != 0:
                log.error("ipptool failed: %s", (perr or pout).strip())
                raise HTTPException(502, "The printer rejected the job.")

        log.info("print user=%s printer=%s color=%s ok", user, PRINTER_URI, color_mode)
        return JSONResponse({"status": "queued", "detail": "Sent to the printer."})
    finally:
        shutil.rmtree(work, ignore_errors=True)
